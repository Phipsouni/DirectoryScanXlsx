import os
import re
import sys
import subprocess

# Директория скрипта (для path.txt и requirements.txt)
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
os.chdir(SCRIPT_DIR)

# --- Установка библиотек при первом запуске ---
def ensure_dependencies():
    try:
        import pandas as pd  # noqa: F401
        import openpyxl  # noqa: F401
        import colorama  # noqa: F401
    except ImportError:
        print("Установка зависимостей при первом запуске...")
        req_path = os.path.join(SCRIPT_DIR, "requirements.txt")
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "-r", req_path],
            cwd=SCRIPT_DIR,
        )
        print("Зависимости установлены. Перезапустите скрипт.")
        sys.exit(0)

ensure_dependencies()

import pandas as pd
import openpyxl
import time
import colorama

colorama.init(autoreset=True)
PATH_FILE = os.path.join(SCRIPT_DIR, "path.txt")

# Цвета для путей в меню
GREEN = colorama.Fore.GREEN
BLUE = colorama.Fore.BLUE
RESET = colorama.Style.RESET_ALL


def normalize_path(s):
    """Убирает внешние кавычки \" и ' из введённого пути."""
    if not s:
        return s
    s = s.strip()
    if (s.startswith('"') and s.endswith('"')) or (s.startswith("'") and s.endswith("'")):
        return s[1:-1].strip()
    return s


def read_saved_paths():
    """Читает сохранённые пути из path.txt. Возвращает (source, save) или (None, None)."""
    if not os.path.isfile(PATH_FILE):
        return None, None
    try:
        with open(PATH_FILE, "r", encoding="utf-8") as f:
            lines = [line.strip() for line in f.readlines() if line.strip()]
        if len(lines) >= 2:
            return lines[0], lines[1]
    except Exception:
        pass
    return None, None


def save_paths(source_path, save_path):
    """Сохраняет пути в path.txt."""
    with open(PATH_FILE, "w", encoding="utf-8") as f:
        f.write(source_path + "\n")
        f.write(save_path + "\n")


def ask_path(prompt, default=None):
    """Запрашивает путь (кавычки при вводе допустимы). Enter — использовать default."""
    raw = input(f"{prompt}: " if not default else f"{prompt} [Enter = текущий]: ")
    s = normalize_path(raw)
    if default and not s:
        return default
    return s


def validate_path(path, kind="директория"):
    if not path:
        return False, "Путь не указан."
    if not os.path.isdir(path):
        return False, f"Директория не найдена: {path}"
    return True, None


def show_menu(source_path, save_path):
    """Выводит меню с текущими путями разными цветами."""
    print()
    print("--- Текущие пути ---")
    src_text = source_path if source_path else "(не задан)"
    sav_text = save_path if save_path else "(не задан)"
    print(f"  Рабочая директория (инвойсы): {GREEN}{src_text}{RESET}")
    print(f"  Директория сохранения файла: {BLUE}{sav_text}{RESET}")
    print()
    print("1. Запуск скрипта")
    print("2. Изменить путь к директории с инвойсами")
    print("3. Изменить путь к директории сохранения файла")
    print("0. Выход")
    print()


def run_scan(source_path, save_path):
    """Выполняет сканирование и создание Excel."""
    folder_number_pattern = re.compile(r"^(\d+)")
    app_name_pattern = re.compile(r"^[^,]+, [^,]+, [^,]+, ([^,]+),")
    esd_pattern = re.compile(r"^([\w-]+)\.pdf$", re.IGNORECASE)
    gtd_pattern = re.compile(r"^GTD_(\d+)_(\d+)_(\d+)\.pdf$", re.IGNORECASE)
    data = []

    for folder_name in os.listdir(source_path):
        folder_path = os.path.join(source_path, folder_name)
        if not os.path.isdir(folder_path):
            continue
        folder_number_match = folder_number_pattern.search(folder_name)
        app_name_match = app_name_pattern.search(folder_name)
        if not folder_number_match or not app_name_match:
            continue
        folder_number = int(folder_number_match.group(1))
        app_name = app_name_match.group(1).strip()
        esd_numbers = []
        gtd_numbers = []
        for file_name in os.listdir(folder_path):
            name_without_ext = file_name[:-4] if file_name.lower().endswith(".pdf") else ""
            if (
                esd_pattern.match(file_name)
                and not file_name.startswith("GTD_")
                and name_without_ext.count("-") == 4
            ):
                esd_numbers.append(file_name[:-4])
            gtd_match = gtd_pattern.match(file_name)
            if gtd_match:
                gtd_numbers.append(
                    f"{gtd_match.group(1)}/{gtd_match.group(2)}/{gtd_match.group(3)}"
                )
        data.append(
            [app_name, folder_number, ", ".join(esd_numbers), ", ".join(gtd_numbers)]
        )

    df = pd.DataFrame(
        data, columns=["Application", "Folder Number", "ЭСД Number", "GTD Number"]
    )
    df_sorted = df.sort_values(by=["Application", "Folder Number"])
    output_data = []
    current_app = None
    for _, row in df_sorted.iterrows():
        app_name, folder_number, esd, gtd = row
        if app_name != current_app:
            output_data.append([app_name, None, None, None])
            current_app = app_name
        output_data.append([None, folder_number, esd, gtd])

    df_final = pd.DataFrame(
        output_data,
        columns=["Application", "Folder Number", "ЭСД Number", "GTD Number"],
    )
    out_file = os.path.join(save_path, "ESD_DT.xlsx")
    df_final.to_excel(out_file, index=False)

    wb = openpyxl.load_workbook(out_file)
    ws = wb.active
    for col_letter in ("A", "B", "C", "D"):
        col_idx = openpyxl.utils.column_index_from_string(col_letter)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            val = cell.value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 80)
    wb.save(out_file)
    wb.close()

    print("Файл успешно создан.")
    time.sleep(2)


def main():
    source_path, save_path = read_saved_paths()

    while True:
        show_menu(source_path, save_path)
        choice = input("Выберите пункт (0–3): ").strip()

        if choice == "0":
            print("Выход.")
            sys.exit(0)

        if choice == "1":
            if not source_path or not save_path:
                print("Сначала укажите оба пути (пункты 2 и 3).")
                time.sleep(2)
                continue
            ok, err = validate_path(source_path, "рабочая")
            if not ok:
                print(err)
                time.sleep(2)
                continue
            ok, err = validate_path(save_path, "сохранения")
            if not ok:
                print(err)
                time.sleep(2)
                continue
            run_scan(source_path, save_path)
            continue

        if choice == "2":
            new_path = ask_path("Новый путь к директории с инвойсами", source_path)
            new_path = new_path or source_path
            ok, err = validate_path(new_path)
            if not ok:
                print(err)
                time.sleep(2)
                continue
            source_path = new_path
            save_paths(source_path, save_path or "")
            print("Путь обновлён.")
            time.sleep(1)
            continue

        if choice == "3":
            new_path = ask_path("Новый путь к директории сохранения", save_path)
            new_path = new_path or save_path
            ok, err = validate_path(new_path)
            if not ok:
                print(err)
                time.sleep(2)
                continue
            save_path = new_path
            save_paths(source_path or "", save_path)
            print("Путь обновлён.")
            time.sleep(1)
            continue

        print("Неверный пункт. Введите 0, 1, 2 или 3.")
        time.sleep(1)


if __name__ == "__main__":
    main()
